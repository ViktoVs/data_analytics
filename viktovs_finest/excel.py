from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def write_df_to_excel(
        df
        , input_file
        , output_file
        , sheet_name
        , start_row = 1
        , start_col = 1
    ):
    wb = load_workbook(input_file)
    ws = wb[sheet_name]
    
    rows = dataframe_to_rows(df)
    
    for r_idx, row in enumerate(rows, start_row):
        for c_idx, value in enumerate(row, start_col):
            ws.cell(row = r_idx
                    , column = c_idx
                    , value = value
                   )
            
    wb.save(output_file)